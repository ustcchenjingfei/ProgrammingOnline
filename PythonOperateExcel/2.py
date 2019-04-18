import openpyxl

workbook = openpyxl.load_workbook(r'豆瓣TOP250电影.xlsx')  # 打开一个工作簿
# print(type(workbook))
# print(workbook.get_sheet_names())  # 获取工作簿中所有表名（已弃用的函数）
# print(workbook.sheetnames)  # 获取工作簿中所有表名（推荐）
worktable = workbook['Sheet']  # 找到指定名称工作表（找不到会报错）,get_sheet_by_name() 也弃用了，后面用workbook['FishC demo']


# 创建
# new_wt = workbook.create_sheet(index=0, title='FishC demo')  # 创建一个新的工作表
# index指定工作表插入的位置，title指定工作表的名称
# print(workbook.sheetnames)  # ['FishC demo', 'Sheet']


# 删除
# workbook.remove(workbook['FishC demo'])  # 删除一个工作表
# print(workbook.sheetnames)  # ['Sheet']


# c = worktable['A2']
# print(c.row)  # 获取行
# print(c.column)  # 获取列
# print(c.coordinate)  # 获取坐标
# print(c.value)  # 获取对应的内容
# print((c.offset(2, 0)).value)  # 指定偏移量（向下做2行，向右0列）


# AAA列是多少列？？496又是多少列呢？(其实是个26进制，1*26^0+1*26^1+1*26^2=703)
# print(openpyxl.cell.cell.column_index_from_string('AAA'))
# print(openpyxl.cell.cell.get_column_letter(496))


# 访问多个单元格
# for each_movie in worktable['A2':'B10']:
#     for each_cell in each_movie:
#         print(each_cell.value,end=' ')
#     print()


# 在整个工作表，打印A列
# for each_row in worktable.rows:
#     print(each_row[0].value)  #每一行的第0个元素，即A列


# 在工作表指定范围，打印A列
# for each_row in worktable.iter_rows(min_row=2,min_col=1, max_row=4, max_col=2):
#     print(each_row[0].value)


# 拷贝工作表
# new_wt=workbook.copy_worksheet(worktable)
# print(type(new_wt))
# workbook.save(r'豆瓣TOP250电影.xlsx')
