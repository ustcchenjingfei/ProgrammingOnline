import openpyxl

"""
wb = openpyxl.Workbook()  # 创建一个工作簿对象
wt1 = wb.create_sheet(title='表1')
wt2 = wb.create_sheet(title='表2')
wt3 = wb.create_sheet(title='表3')
wt4 = wb.create_sheet(title='表4')

# 个性化工作标签栏
wt1.sheet_properties.tabColor = 'FF0000'  # 红色 （RGB  十六进制）
wt2.sheet_properties.tabColor = '00FF00'
wt3.sheet_properties.tabColor = '0000FF'
wt4.sheet_properties.tabColor = '8B008B'
wb.save(r'change_lable_color.xlsx ')

# 调整宽、高
wt2.row_dimensions[2].height = 100  # 调整第2行高（与列宽单位不一样）
wt2.column_dimensions['C'].width = 50  # 调整第C列宽（行高与单位不一样）
wb.save(r'change_lable_color.xlsx ')

# 合并、拆分单元格
wt1.merge_cells('A1:C3')  # 合并
wt1['A1'] = 'Hello World'
wb.save(r'change_lable_color.xlsx ')

wt1.unmerge_cells('A1:C3')  # 拆分后，内容还在（拆分范围要与合并的范围保持一致，否则报错）
wb.save(r'change_lable_color.xlsx ')
"""

# 冻结窗口
wb = openpyxl.load_workbook(r'豆瓣TOP250电影.xlsx')
wt = wb.active

wt.freeze_panes = 'B2'  # 第2行，第B列(即上边、左边)固定，不包括第2行，第B列。
wb.save('豆瓣TOP250电影.xlsx')

# wt.freeze_panes = 'A1'  # 解冻
wt.freeze_panes = None  # 解冻
wb.save('豆瓣TOP250电影.xlsx')
