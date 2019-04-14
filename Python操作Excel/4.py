from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import GradientFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle

# wb 代表 workbook ,wt 代表 worktable
wb = Workbook()
wt = wb.active
b2 = wt['B2']
b2.value = "FishC"


# 改变单元格字体
bold_red_font = Font(bold=True, color='FF0000')  # 加粗、红色
b2.font = bold_red_font
b3 = wt['B3']
b3.value = "FishC"
italic_strike_blue_16font = Font(size=16, italic=True, color='0000FF', strike=True)
# 字体大小、倾斜、蓝色、删除线
b3.font = italic_strike_blue_16font
wb.save(r'4.py_example.xlsx')


# 填充单元格
yellow_fill = PatternFill(fill_type='solid', fgColor='FFFF00')  # 纯色填充 ,黄色
b2.fill = yellow_fill
wb.save(r'4.py_example.xlsx')


# 从红色到绿色填充，线性填充
red2green_fill = GradientFill(type='linear', stop=('FF0000', '00FF00'))
b3.fill = red2green_fill
wb.save(r'4.py_example.xlsx')


# 设置边框
thin_side = Side(border_style='thin', color='000000') #细线，黑色
double_side = Side(border_style='double', color='FF0000')
b2.border = Border(diagonal=thin_side, diagonalUp=True, diagonalDown=True)  # 对角线
b3.border = Border(left=double_side, top=double_side, right=double_side, bottom=double_side)  # 双红线边框
wb.save(r'4.py_example.xlsx')


# 文本对齐
wt.merge_cells('A1:C2')
wt['A1'].value = 'I love FishC.com'
center_alignment = Alignment(horizontal='center', vertical='center')  # 水平、垂直
wt['A1'].alignment = center_alignment
wb.save(r'4.py_example.xlsx')


# 命名样式
# 设置命名样式
highlight = NamedStyle(name='highlight')
highlight.font = Font(bold=True, size=20)
highlight.alignment = Alignment(horizontal='center', vertical='center')
wb.add_named_style(highlight)  # 注册到工作簿
wt['A1'].style = highlight
wt['B5'].value = 'Love'
wt['B5'].style = highlight
wb.save(r'4.py_example.xlsx')
